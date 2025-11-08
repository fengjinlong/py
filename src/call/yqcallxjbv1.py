import pandas as pd
import matplotlib.pyplot as plt
import re
import numpy as np
from matplotlib.patches import FancyArrowPatch
import os
import glob
import sys

def find_csv_files(data_dir="data"):
    """查找data目录下的所有CSV文件"""
    if not os.path.exists(data_dir):
        print(f"错误: {data_dir} 目录不存在")
        return []
    
    csv_files = glob.glob(os.path.join(data_dir, "*.csv"))
    if not csv_files:
        print(f"错误: {data_dir} 目录下没有找到CSV文件")
        return []
    
    return csv_files

def process_single_file(file_path):
    """处理单个CSV文件"""
    print(f"\n{'='*60}")
    print(f"正在处理文件: {os.path.basename(file_path)}")
    print(f"{'='*60}")
    
    try:
        # 读取 CSV
        df = pd.read_csv(file_path)
        
        # 从文件名提取基础名称（去掉扩展名）
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        
        # 创建 export 文件夹（如果不存在）
        export_dir = "export"
        os.makedirs(export_dir, exist_ok=True)
        
        # 1. 只保留看涨期权（C 结尾）
        df = df[df["产品"].str.endswith("-C")].copy()
        
        if len(df) == 0:
            print(f"警告: {os.path.basename(file_path)} 中没有找到看涨期权数据")
            return None
        
        # 2. 从 "产品" 字段解析出行权价
        # 格式类似：ETH-26DEC25-1200-C
        df["Strike"] = df["产品"].apply(lambda x: int(re.findall(r"-(\d+)-C$", x)[0]))
        
        # 3. 转换关键列为数值型（避免有 "-" 字符）
        # 使用 Δ|增量 列作为 Delta 值
        for col in ["Δ|增量", "Theta", "Gamma", "Vega"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        
        # 去掉缺失值
        df = df.dropna(subset=["Δ|增量", "Theta", "Vega"])
        
        # 添加Theta过滤条件：剔除Theta绝对值太小的点
        df = df[df["Theta"].abs() >= 1e-3].copy()
        
        # 4. 计算性价比指标
        df["Delta/Theta"] = df["Δ|增量"] / df["Theta"].abs()
        df["Gamma/Theta"] = df["Gamma"] / df["Theta"].abs()
        df["Vega/Theta"] = df["Vega"] / df["Theta"].abs()
        
        # 5. 初始化推荐列
        df["Recommendation"] = "Normal"
        
        # 6. 定义OTM筛选条件：Delta范围筛选
        delta_min = 0.15 # 最小Delta值
        delta_max = 0.45   # 最大Delta值
        otm_condition = (df["Δ|增量"].abs() >= delta_min) & (df["Δ|增量"].abs() <= delta_max)
        
        print(f"OTM筛选条件: {delta_min} ≤ |Delta| ≤ {delta_max}")
        print(f"符合OTM条件的合约数量: {otm_condition.sum()}")
        print(f"不符合OTM条件的合约数量: {(~otm_condition).sum()}")
        
        # 标记前 3 名 (Delta/Theta) - 仅在OTM范围内筛选
        otm_df = df[otm_condition].copy()
        if len(otm_df) >= 3:
            top3_delta = otm_df["Delta/Theta"].nlargest(3).index
            for rank, idx in enumerate(top3_delta, start=1):
                df.loc[idx, "Recommendation"] = f"Top{rank} (Delta/Theta)"
            print(f"\n前3名Delta/Theta (OTM范围): {len(top3_delta)}个")
        else:
            top3_delta = []
            print(f"\n警告: OTM范围内合约数量不足3个 ({len(otm_df)}个)")
        
        # 标记前 3 名 (Gamma/Theta) - 不限制OTM条件
        top3_gamma = df["Gamma/Theta"].nlargest(3).index
        for rank, idx in enumerate(top3_gamma, start=1):
            if df.loc[idx, "Recommendation"] == "Normal":
                df.loc[idx, "Recommendation"] = f"Top{rank} (Gamma/Theta)"
            else:
                df.loc[idx, "Recommendation"] += f" + Top{rank} (Gamma/Theta)"
        
        # 标记前 3 名 (Vega/Theta) - 仅在OTM范围内筛选
        if len(otm_df) >= 3:
            top3_vega = otm_df["Vega/Theta"].nlargest(3).index
            for rank, idx in enumerate(top3_vega, start=1):
                if df.loc[idx, "Recommendation"] == "Normal":
                    df.loc[idx, "Recommendation"] = f"Top{rank} (Vega/Theta)"
                else:
                    df.loc[idx, "Recommendation"] += f" + Top{rank} (Vega/Theta)"
            print(f"\n前3名Vega/Theta (OTM范围): {len(top3_vega)}个")
        else:
            top3_vega = []
            print(f"\n警告: OTM范围内合约数量不足3个，无法筛选Vega/Theta")
        
        # 7. 打印推荐结果
        if len(top3_delta) > 0:
            print("\n前 3 名 Delta/Theta 行权价 (OTM范围):")
            print(df.loc[top3_delta, ["产品", "Strike", "Δ|增量", "Delta/Theta"]])
        else:
            print("\n前 3 名 Delta/Theta 行权价 (OTM范围): 无符合条件的数据")
        
        print("\n前 3 名 Gamma/Theta 行权价:")
        print(df.loc[top3_gamma, ["产品", "Strike", "Gamma", "Gamma/Theta"]])
        
        if len(top3_vega) > 0:
            print("\n前 3 名 Vega/Theta 行权价 (OTM范围):")
            print(df.loc[top3_vega, ["产品", "Strike", "Vega", "Vega/Theta"]])
        else:
            print("\n前 3 名 Vega/Theta 行权价 (OTM范围): 无符合条件的数据")
        
        # 8. 生成图表
        generate_charts(df, otm_df, top3_delta, top3_gamma, top3_vega, otm_condition, base_name, export_dir)
        
        # 9. 生成Excel和CSV文件
        generate_output_files(df, top3_delta, top3_gamma, top3_vega, base_name, export_dir)
        
        # 10. 打印统计信息
        print_statistics(file_path, df, otm_condition, delta_min, delta_max, base_name, export_dir)
        
        return df
        
    except Exception as e:
        print(f"处理文件 {os.path.basename(file_path)} 时出错: {str(e)}")
        return None

def generate_charts(df, otm_df, top3_delta, top3_gamma, top3_vega, otm_condition, base_name, export_dir):
    """生成图表"""
    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False
    
    # 画图 - 现在有3个子图
    fig = plt.figure(figsize=(24, 10))

    # 添加总标题，显示文件名
    fig.suptitle(f'期权分析报告 - {base_name}', fontsize=16, fontweight='bold', y=0.95)
    
    # Δ/|Θ| 曲线 - 使用对数坐标
    plt.subplot(1, 3, 1)
    plt.plot(df["Strike"], df["Delta/Theta"], marker="o", alpha=0.7, label="All", color="lightgray", markersize=4)
    plt.plot(otm_df["Strike"], otm_df["Delta/Theta"], marker="o", color="blue", 
             linewidth=2.5, label="OTM Range", alpha=0.9, markersize=6)
    
    if len(top3_delta) > 0:
        plt.scatter(df.loc[top3_delta, "Strike"], df.loc[top3_delta, "Delta/Theta"],
                    color="green", s=120, label="Top3 Delta/Theta", zorder=5, alpha=0.9, 
                    edgecolors='darkgreen', linewidth=2)
        
        # 使用箭头指引，将标签放在空白区域
        top3_delta_sorted = df.loc[top3_delta].sort_values("Strike")
        for i, (idx, row) in enumerate(top3_delta_sorted.iterrows()):
            # 计算标签位置 - 放在空白区域
            strike = row["Strike"]
            value = row["Delta/Theta"]
            
            # 根据位置选择标签位置
            if i == 0:  # 第一个点 - 放在左上方
                label_x = strike - 200
                label_y = value * 4
                arrow_start = (strike - 50, value * 1.1)
            elif i == 1:  # 第二个点 - 放在右上方
                label_x = strike + 200
                label_y = value * 8
                arrow_start = (strike + 50, value * 1.1)
            else:  # 第三个点 - 放在右下方
                label_x = strike + 200
                label_y = value * 16
                arrow_start = (strike + 50, value * 0.9)
            
            # 添加箭头
            arrow = FancyArrowPatch((arrow_start[0], arrow_start[1]), 
                                   (label_x, label_y),
                                   arrowstyle='->', mutation_scale=15, 
                                   color='green', linewidth=2, alpha=0.8)
            plt.gca().add_patch(arrow)
            
            # 添加标签
            plt.text(label_x, label_y, str(int(strike)), 
                     fontsize=10, ha="center", va="center",
                     color="green", fontweight="bold",
                     bbox=dict(boxstyle="round,pad=0.3", facecolor="white", 
                              edgecolor="green", alpha=0.9, linewidth=2))
    
    # 添加OTM范围标记线
    plt.axhline(y=df[otm_condition]["Delta/Theta"].max(), color="red", linestyle="--", alpha=0.5, label="OTM Max")
    plt.axhline(y=df[otm_condition]["Delta/Theta"].min(), color="red", linestyle="--", alpha=0.5, label="OTM Min")
    
    plt.title("Delta/Theta vs Strike (Log Scale)\nOTM Range Highlighted", fontsize=12, fontweight="bold")
    plt.xlabel("Strike Price", fontsize=11, fontweight="bold")
    plt.ylabel("Delta/|Theta| (Log Scale)", fontsize=11, fontweight="bold")
    plt.yscale('log')  # 使用对数坐标
    plt.legend(fontsize=9)
    plt.grid(True, alpha=0.3)
    
    # Γ/|Θ| 曲线
    plt.subplot(1, 3, 2)
    plt.plot(df["Strike"], df["Gamma/Theta"], marker="o", alpha=0.7, color="lightgray", markersize=4)
    plt.scatter(df.loc[top3_gamma, "Strike"], df.loc[top3_gamma, "Gamma/Theta"],
                color="green", s=120, label="Top3 Gamma/Theta", zorder=5, alpha=0.9, 
                edgecolors='darkgreen', linewidth=2)
    
    # 使用箭头指引，将标签放在空白区域
    top3_gamma_sorted = df.loc[top3_gamma].sort_values("Strike")
    for i, (idx, row) in enumerate(top3_gamma_sorted.iterrows()):
        # 计算标签位置 - 放在空白区域
        strike = row["Strike"]
        value = row["Gamma/Theta"]
        
        # 根据位置选择标签位置
        if i == 0:  # 第一个点 - 放在左上方
            label_x = strike - 300
            label_y = value * 1.2
            arrow_start = (strike - 100, value * 1.05)
        elif i == 1:  # 第二个点 - 放在右上方
            label_x = strike + 300
            label_y = value * 1.2
            arrow_start = (strike + 100, value * 1.05)
        else:  # 第三个点 - 放在右下方
            label_x = strike + 300
            label_y = value * 0.8
            arrow_start = (strike + 100, value * 0.95)
        
        # 添加箭头
        arrow = FancyArrowPatch((arrow_start[0], arrow_start[1]), 
                               (label_x, label_y),
                               arrowstyle='->', mutation_scale=15, 
                               color='green', linewidth=2, alpha=0.8)
        plt.gca().add_patch(arrow)
        
        # 添加标签
        plt.text(label_x, label_y, str(int(strike)), 
                 fontsize=10, ha="center", va="center",
                 color="green", fontweight="bold",
                 bbox=dict(boxstyle="round,pad=0.3", facecolor="white", 
                          edgecolor="green", alpha=0.9, linewidth=2))
    
    plt.title("Gamma/Theta vs Strike", fontsize=12, fontweight="bold")
    plt.xlabel("Strike Price", fontsize=11, fontweight="bold")
    plt.ylabel("Gamma/|Theta|", fontsize=11, fontweight="bold")
    plt.legend(fontsize=9)
    plt.grid(True, alpha=0.3)
    
    # Vega/|Θ| 曲线 - 大幅优化第三个子图
    plt.subplot(1, 3, 3)
    
    # 绘制所有数据点（浅色）
    plt.plot(df["Strike"], df["Vega/Theta"], marker="o", alpha=0.4, color="lightgray", 
             markersize=3, linewidth=1, label="All Contracts")
    
    # 绘制OTM范围数据点（蓝色，更突出）
    plt.plot(otm_df["Strike"], otm_df["Vega/Theta"], marker="o", color="steelblue", 
             linewidth=3, label="OTM Range", alpha=0.8, markersize=8, 
             markerfacecolor="lightblue", markeredgecolor="steelblue", markeredgewidth=2)
    
    # 绘制推荐点（橙色，最突出）
    if len(top3_vega) > 0:
        plt.scatter(df.loc[top3_vega, "Strike"], df.loc[top3_vega, "Vega/Theta"],
                    color="orange", s=150, label="Top3 Vega/Theta", zorder=10, alpha=0.95, 
                    edgecolors='darkorange', linewidth=3, marker='D')
        
        # 优化箭头指引和标签
        top3_vega_sorted = df.loc[top3_vega].sort_values("Strike")
        for i, (idx, row) in enumerate(top3_vega_sorted.iterrows()):
            strike = row["Strike"]
            value = row["Vega/Theta"]
            
            # 获取Strike的范围来动态调整标签位置
            strike_min = df["Strike"].min()
            strike_max = df["Strike"].max()
            strike_range = strike_max - strike_min
            
            # 获取Vega/Theta的范围来动态调整标签位置
            vega_min = df["Vega/Theta"].min()
            vega_max = df["Vega/Theta"].max()
            vega_range = vega_max - vega_min
            
            # 根据位置选择标签位置，避免重叠
            if i == 0:  # 第一个点 - 放在左上方
                label_x = strike - strike_range * 0.12
                label_y = value + vega_range * 0.08
                arrow_start = (strike - strike_range * 0.04, value + vega_range * 0.02)
            elif i == 1:  # 第二个点 - 放在右上方
                label_x = strike + strike_range * 0.12
                label_y = value + vega_range * 0.08
                arrow_start = (strike + strike_range * 0.04, value + vega_range * 0.02)
            else:  # 第三个点 - 放在右下方
                label_x = strike + strike_range * 0.12
                label_y = value - vega_range * 0.08
                arrow_start = (strike + strike_range * 0.04, value - vega_range * 0.02)
            
            # 添加箭头（更粗更明显）
            arrow = FancyArrowPatch((arrow_start[0], arrow_start[1]), 
                                   (label_x, label_y),
                                   arrowstyle='->', mutation_scale=25, 
                                   color='darkorange', linewidth=3, alpha=0.9)
            plt.gca().add_patch(arrow)
            
            # 添加标签（更大更明显）
            plt.text(label_x, label_y, str(int(strike)), 
                     fontsize=12, ha="center", va="center",
                     color="darkorange", fontweight="bold",
                     bbox=dict(boxstyle="round,pad=0.5", facecolor="white", 
                              edgecolor="darkorange", alpha=0.95, linewidth=3))
    
    # 添加OTM范围标记线（更明显）
    otm_vega_max = df[otm_condition]["Vega/Theta"].max()
    otm_vega_min = df[otm_condition]["Vega/Theta"].min()
    plt.axhline(y=otm_vega_max, color="red", linestyle="--", alpha=0.7, 
                label="OTM Max", linewidth=2.5)
    plt.axhline(y=otm_vega_min, color="red", linestyle="--", alpha=0.7, 
                label="OTM Min", linewidth=2.5)
    
    # 添加OTM范围填充区域（半透明红色背景）
    plt.fill_between(otm_df["Strike"], otm_vega_min, otm_vega_max, 
                     alpha=0.15, color="red", label="OTM Zone")
    
    # 添加数值标注（在OTM范围线上）
    plt.text(otm_df["Strike"].mean(), otm_vega_max + vega_range * 0.02, 
             f"Max: {otm_vega_max:.2f}", ha="center", va="bottom", 
             fontsize=9, color="red", fontweight="bold")
    plt.text(otm_df["Strike"].mean(), otm_vega_min - vega_range * 0.02, 
             f"Min: {otm_vega_min:.2f}", ha="center", va="top", 
             fontsize=9, color="red", fontweight="bold")
    
    # 设置标题和标签（更突出）
    plt.title("Vega/Theta vs Strike\nOTM Range Highlighted", fontsize=13, fontweight="bold", pad=20)
    plt.xlabel("Strike Price", fontsize=12, fontweight="bold")
    plt.ylabel("Vega/|Theta|", fontsize=12, fontweight="bold")
    
    # 优化图例
    plt.legend(loc='upper right', fontsize=9, framealpha=0.9, 
               fancybox=True, shadow=True)
    
    # 设置网格
    plt.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    
    # 设置y轴范围，确保标签可见
    if len(top3_vega) > 0:
        y_margin = vega_range * 0.15
        y_max = max(df["Vega/Theta"].max(), max([df.loc[idx, "Vega/Theta"] + y_margin for idx in top3_vega]))
        y_min = min(df["Vega/Theta"].min(), min([df.loc[idx, "Vega/Theta"] - y_margin for idx in top3_vega]))
        plt.ylim(y_min, y_max)
    
    # 添加统计信息文本框
    stats_text = f"OTM Contracts: {len(otm_df)}\nTop3 Vega/Theta: {len(top3_vega)}"
    plt.text(0.02, 0.98, stats_text, transform=plt.gca().transAxes, 
             fontsize=9, verticalalignment='top', 
             bbox=dict(boxstyle="round,pad=0.5", facecolor="lightblue", alpha=0.8))

    # 添加文件信息文本框
    file_info = f"数据文件: {base_name}.csv\n生成时间: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
    plt.text(0.98, 0.98, file_info, transform=fig.transFigure,
             fontsize=10, verticalalignment='top', horizontalalignment='right',
             bbox=dict(boxstyle="round,pad=0.5", facecolor="lightyellow", alpha=0.9))
    
    plt.tight_layout()
    
    # 根据输入文件名生成输出文件名，保存到 export 文件夹
    image_file = os.path.join(export_dir, f"{base_name}_options_analysis.png")
    plt.savefig(image_file, dpi=300, bbox_inches='tight', facecolor='white')
    print(f"\n图片已保存为: {image_file}")
    
    plt.show()

def generate_output_files(df, top3_delta, top3_gamma, top3_vega, base_name, export_dir):
    """生成Excel和CSV输出文件"""
    # 创建带不同颜色标记的Excel文件
    output_file = os.path.join(export_dir, f"{base_name}_options_with_recommendation.xlsx")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 写入数据
        df.to_excel(writer, sheet_name='期权分析', index=False)
        
        # 获取工作表
        worksheet = writer.sheets['期权分析']
        
        # 导入样式
        from openpyxl.styles import PatternFill, Font, Border, Side
        
        # 定义不同颜色背景
        delta_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 浅绿色 - Delta/Theta
        gamma_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # 天蓝色 - Gamma/Theta
        vega_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")   # 粉色 - Vega/Theta
        both_fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")   # 紫色 - 两者都有
        triple_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # 金色 - 三者都有
        bold_font = Font(bold=True)
        
        # 定义边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 为前3名Delta/Theta添加浅绿色背景
        for idx in top3_delta:
            row_num = df.index.get_loc(idx) + 2  # +2 因为Excel从1开始，还有标题行
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = delta_fill
                cell.font = bold_font
                cell.border = thin_border
        
        # 为前3名Gamma/Theta添加天蓝色背景
        for idx in top3_gamma:
            row_num = df.index.get_loc(idx) + 2  # +2 因为Excel从1开始，还有标题行
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                # 检查是否已经有其他标记
                if "Delta/Theta" in df.loc[idx, "Recommendation"] and "Vega/Theta" in df.loc[idx, "Recommendation"]:
                    cell.fill = triple_fill  # 金色 - 三者都有
                elif "Delta/Theta" in df.loc[idx, "Recommendation"] or "Vega/Theta" in df.loc[idx, "Recommendation"]:
                    cell.fill = both_fill  # 紫色 - 两者都有
                else:
                    cell.fill = gamma_fill  # 天蓝色 - 只有Gamma/Theta
                cell.font = bold_font
                cell.border = thin_border
        
        # 为前3名Vega/Theta添加粉色背景
        for idx in top3_vega:
            row_num = df.index.get_loc(idx) + 2  # +2 因为Excel从1开始，还有标题行
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                # 检查是否已经有其他标记
                if "Delta/Theta" in df.loc[idx, "Recommendation"] and "Gamma/Theta" in df.loc[idx, "Recommendation"]:
                    cell.fill = triple_fill  # 金色 - 三者都有
                elif "Delta/Theta" in df.loc[idx, "Recommendation"] or "Vega/Theta" in df.loc[idx, "Recommendation"]:
                    cell.fill = both_fill  # 紫色 - 两者都有
                else:
                    cell.fill = vega_fill  # 粉色 - 只有Vega/Theta
                cell.font = bold_font
                cell.border = thin_border
        
        # 为所有单元格添加边框
        for row in range(1, len(df) + 2):  # +2 因为包含标题行
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                if not cell.border.left:  # 如果还没有边框
                    cell.border = thin_border
    
    # 创建增强版CSV文件（添加颜色标记说明）
    csv_file = os.path.join(export_dir, f"{base_name}_options_with_recommendation.csv")
    
    # 创建带颜色标记的CSV数据
    csv_df = df.copy()
    
    # 添加颜色标记列
    def get_color_mark(recommendation):
        if "Delta/Theta" in recommendation and "Gamma/Theta" in recommendation and "Vega/Theta" in recommendation:
            return "🟡 金色标记 (Delta+Gamma+Vega三优)"
        elif "Delta/Theta" in recommendation and "Gamma/Theta" in recommendation:
            return "🟣 紫色标记 (Delta+Gamma双优)"
        elif "Delta/Theta" in recommendation and "Vega/Theta" in recommendation:
            return "🟣 紫色标记 (Delta+Vega双优)"
        elif "Gamma/Theta" in recommendation and "Vega/Theta" in recommendation:
            return "🟣 紫色标记 (Gamma+Vega双优)"
        elif "Delta/Theta" in recommendation:
            return "🟢 绿色标记 (Delta/Theta优)"
        elif "Gamma/Theta" in recommendation:
            return "🔵 蓝色标记 (Gamma/Theta优)"
        elif "Vega/Theta" in recommendation:
            return "🩷 粉色标记 (Vega/Theta优)"
        else:
            return "⚪ 普通"
    
    csv_df["颜色标记"] = csv_df["Recommendation"].apply(get_color_mark)
    
    # 重新排列列，把颜色标记放在前面
    cols = ["颜色标记", "产品", "Strike", "Δ|增量", "Gamma", "Vega", "Theta", 
            "Delta/Theta", "Gamma/Theta", "Vega/Theta", "Recommendation"]
    csv_df = csv_df[cols]
    
    # 保存CSV
    csv_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
    
    print(f"\n结果已写入 {output_file} (带颜色标记)")
    print(f"结果已写入 {csv_file} (带颜色说明)")

def print_statistics(file_path, df, otm_condition, delta_min, delta_max, base_name, export_dir):
    """打印统计信息"""
    print(f"\n过滤前数据点数量: {len(pd.read_csv(file_path)[pd.read_csv(file_path)['产品'].str.endswith('-C')])}")
    print(f"过滤后数据点数量: {len(df)}")
    print(f"Theta过滤条件: |Theta| >= 1e-3")
    print(f"OTM筛选条件: {delta_min} ≤ |Delta| ≤ {delta_max}")
    print(f"符合OTM条件的合约: {otm_condition.sum()}个")
    
    # 打印颜色标记说明
    print(f"\n=== 颜色标记说明 ===")
    print(f"🟢 绿色标记: Top3 Delta/Theta (OTM范围内)")
    print(f"🔵 蓝色标记: Top3 Gamma/Theta (全范围)")
    print(f"🩷 粉色标记: Top3 Vega/Theta (OTM范围内)")
    print(f"🟣 紫色标记: 同时获得两种推荐")
    print(f"🟡 金色标记: 同时获得三种推荐")
    print(f"⚪ 普通标记: 未获得推荐")
    
    # 打印文件命名信息
    print(f"\n=== 文件命名规则 ===")
    print(f"输入文件: {os.path.basename(file_path)}")
    print(f"输出图片: {export_dir}/{base_name}_options_analysis.png")
    print(f"输出Excel: {export_dir}/{base_name}_options_with_recommendation.xlsx")
    print(f"输出CSV: {export_dir}/{base_name}_options_with_recommendation.csv")
    print(f"基础名称: {base_name}")
    print(f"输出目录: {export_dir}")

def main():
    """主函数"""
    print("期权分析工具 - 优化版")
    print("=" * 50)
    
    # 查找CSV文件
    csv_files = find_csv_files("data")
    
    if not csv_files:
        print("没有找到CSV文件，程序退出")
        return
    
    print(f"找到 {len(csv_files)} 个CSV文件:")
    for i, file_path in enumerate(csv_files, 1):
        print(f"  {i}. {os.path.basename(file_path)}")
    
    # 处理每个文件
    processed_count = 0
    for file_path in csv_files:
        result = process_single_file(file_path)
        if result is not None:
            processed_count += 1
    
    print(f"\n{'='*60}")
    print(f"处理完成！成功处理 {processed_count}/{len(csv_files)} 个文件")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()
